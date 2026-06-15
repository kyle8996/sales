"""
从 Excel 重新导入跟进记录 + 更新报名日期
"""
import openpyxl, requests, re
from datetime import datetime

SUPABASE_URL = "https://chasxggorljjqqmficnh.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNoYXN4Z2dvcmxqanFxbWZpY25oIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODE1MDI1NzIsImV4cCI6MjA5NzA3ODU3Mn0.E0xlNCX_jFn8X9LheAR3jFHokzWSkZlXkBDReqTvL6c"
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def parse_date(raw):
    """从咨询日期/Excel日期解析为 YYYY-MM-DD 字符串"""
    if not raw or raw == 'None':
        return None
    if isinstance(raw, datetime):
        return raw.strftime('%Y-%m-%d')
    raw = str(raw).strip()
    if raw == '年前':
        return None  # 无法确定具体日期
    # 尝试解析 2026-03-02 格式
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', raw)
    if m:
        return f"{int(m.group(1))}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r'(\d{4})/(\d{1,2})/(\d{1,2})', raw)
    if m:
        return f"{int(m.group(1))}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r'(\d{1,2})\.(\d{1,2})', raw)
    if m:
        return f"2026-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None

def map_status(raw):
    if not raw:
        return "未上门"
    raw = raw.strip()
    if "报名" in raw or "已报" in raw or "已交" in raw or "在读" in raw or "学员" in raw:
        return "已报名"
    if "上门" in raw or "未缴" in raw or "试听" in raw:
        return "已上门未报名"
    return "未上门"

def main():
    path = "C:/Users/刘新/Desktop/DK邻里少儿中心-客户信息登记表.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["私域客户登记表"]

    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[col_idx] = str(cell.value).strip()

    def get(row, col_name):
        for idx, h in headers.items():
            if col_name in h:
                val = row[idx - 1]
                return val
        return None

    # 先获取现有的 customers（用于匹配名字）
    print("📡 获取现有客户列表...")
    r = requests.get(f"{SUPABASE_URL}/rest/v1/customers?select=id,name,status,notes", headers={**HEADERS, "Accept": "application/json"})
    existing_customers = r.json()
    print(f"  现有 {len(existing_customers)} 条客户记录")

    # 构建名字到ID的映射
    name_to_ids = {}
    for c in existing_customers:
        n = c['name']
        if n not in name_to_ids:
            name_to_ids[n] = []
        name_to_ids[n].append(c)

    # 清理旧跟进记录
    print("🗑 清理旧跟进记录...")
    requests.delete(f"{SUPABASE_URL}/rest/v1/followups?id=gt.0", headers=HEADERS)

    followup_count = 0
    enrollment_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = get(row, "学员姓名")
        if not name or str(name).strip() == 'None':
            continue
        name = str(name).strip()

        notes = get(row, "咨询情况记录")
        notes_str = str(notes).strip() if notes else None

        consult_date = get(row, "咨询日期")
        status_raw = get(row, "客户状态")
        status = map_status(str(status_raw).strip() if status_raw else None)
        owner = get(row, "跟进人")
        owner_str = str(owner).strip() if owner else None

        # 找到匹配的客户
        matches = name_to_ids.get(name, [])
        if not matches:
            print(f"  ⚠️ 未找到客户: {name}")
            continue
        
        cust = matches[0]  # 取第一个匹配
        cid = cust['id']

        # 1) 创建跟进记录 (从咨询情况记录)
        if notes_str:
            fu_data = {
                "customer_id": cid,
                "content": notes_str,
                "need_reminder": False,
                "trial_invited": False,
                "owner": owner_str or "Kyle",
                "created_at": datetime.now().isoformat()
            }
            resp = requests.post(f"{SUPABASE_URL}/rest/v1/followups", headers=HEADERS, json=fu_data)
            if resp.status_code in [200, 201, 204]:
                followup_count += 1

        # 2) 更新报名日期 (仅已报名客户)
        if status == "已报名":
            enroll_date = parse_date(consult_date)
            if enroll_date:
                resp = requests.patch(
                    f"{SUPABASE_URL}/rest/v1/customers?id=eq.{cid}",
                    headers=HEADERS,
                    json={"enrollment_date": enroll_date}
                )
                if resp.status_code in [200, 201, 204]:
                    enrollment_count += 1

    print(f"\n🎉 完成!")
    print(f"  创建跟进记录: {followup_count} 条")
    print(f"  更新报名日期: {enrollment_count} 条")

if __name__ == "__main__":
    main()
