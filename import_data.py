"""
导入「DK邻里少儿中心-客户信息登记表.xlsx」私域客户登记表到 Supabase
"""
import openpyxl, requests, json, re
from datetime import datetime

# Supabase config
SUPABASE_URL = "https://chasxggorljjqqmficnh.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNoYXN4Z2dvcmxqanFxbWZpY25oIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODE1MDI1NzIsImV4cCI6MjA5NzA3ODU3Mn0.E0xlNCX_jFn8X9LheAR3jFHokzWSkZlXkBDReqTvL6c"
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def map_status(raw):
    """映射客户状态到标准化状态"""
    if not raw:
        return "新线索"
    raw = raw.strip()
    if "报名" in raw or "已报" in raw or "已交" in raw:
        return "已成交"
    if "上门" in raw or "未缴" in raw or "试听" in raw:
        return "已试听"
    if "邀约" in raw or "预约" in raw:
        return "已邀约"
    if "联系" in raw:
        return "已联系"
    if "流失" in raw or "拒绝" in raw or "不要" in raw:
        return "已流失"
    if "在读" in raw or "学员" in raw:
        return "已成交"
    return "新线索"

def parse_trial_date(raw):
    """解析体验课邀约时间"""
    if not raw:
        return None
    raw = str(raw).strip()
    # 尝试各种格式
    patterns = [
        (r'(\d+)\.(\d+).*?(\d+):(\d+)', 'M.D H:M'),  # 3.11周三19:30
        (r'(\d+)/(\d+).*?(\d+):(\d+)', 'M/D H:M'),  # 3/11 19:30
        (r'(\d+)-(\d+).*?(\d+):(\d+)', 'M-D H:M'),  # 3-11 19:30
    ]
    for pat, fmt in patterns:
        m = re.search(pat, raw)
        if m:
            month, day, hour, minute = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
            year = 2026
            try:
                dt = datetime(year, month, day, hour, minute)
                return dt.isoformat()
            except:
                return None
    return None

def parse_parent(raw):
    """从原始字段提取家长称呼"""
    if not raw:
        return None
    raw = str(raw).strip()
    # 常见: 妈妈, 爸爸, 奶奶, 外婆...
    parents = ["妈妈", "爸爸", "奶奶", "外婆", "外公", "爷爷", "姑姑", "阿姨", "叔叔"]
    for p in parents:
        if p in raw:
            return p
    return raw[:5]  # 截取前5字

def main():
    # 读取 Excel
    path = "C:/Users/刘新/Desktop/DK邻里少儿中心-客户信息登记表.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["私域客户登记表"]

    # 解析字段映射 (Row 1 = headers)
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[col_idx] = str(cell.value).strip()

    print("📋 Excel 表头映射:")
    for k, v in headers.items():
        print(f"  Col {k}: {v}")

    # 解析数据行 (Row 2+)
    customers = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # 获取各字段
        def get(col_name):
            for idx, h in headers.items():
                if col_name in h:
                    val = row[idx - 1]
                    return str(val).strip() if val else None
            return None

        name = get("学员姓名")
        if not name or name == "None":
            skipped += 1
            continue

        source = get("客户来源")
        status_raw = get("客户状态")
        notes = get("咨询情况记录")
        owner = get("跟进人")
        trial_raw = get("体验课邀约时间")
        parent_raw = get("家长")
        phone = get("联系电话")
        consult_date_raw = get("咨询日期")

        customer = {
            "name": name,
            "parent": parse_parent(parent_raw) or parent_raw,
            "phone": phone,
            "age": get("年龄"),
            "grade": get("年级"),
            "school": get("学校"),
            "source": source,
            "status": map_status(status_raw),
            "owner": owner,
            "notes": notes,
            "trial_date": parse_trial_date(trial_raw),
            "tags": get("备注"),
        }
        customers.append(customer)

    print(f"\n📊 共解析 {len(customers)} 条客户记录，跳过 {skipped} 行")

    # 统计状态分布
    status_count = {}
    for c in customers:
        s = c["status"]
        status_count[s] = status_count.get(s, 0) + 1
    print("\n📈 状态分布:")
    for s, n in sorted(status_count.items(), key=lambda x: x[1], reverse=True):
        print(f"  {s}: {n}人")

    # 批量导入到 Supabase（每次最多 50 条）
    # 确保所有记录有相同字段（Supabase 要求）
    all_keys = ["name", "parent", "phone", "age", "grade", "school", "source",
                "status", "owner", "notes", "trial_date", "tags"]
    batch_size = 50
    total_imported = 0
    for i in range(0, len(customers), batch_size):
        batch = customers[i:i+batch_size]
        clean_batch = []
        for c in batch:
            clean = {}
            for k in all_keys:
                clean[k] = c.get(k)
            clean_batch.append(clean)

        url = f"{SUPABASE_URL}/rest/v1/customers"
        resp = requests.post(url, headers=HEADERS, json=clean_batch)
        if resp.status_code in [200, 201, 204]:
            total_imported += len(clean_batch)
            print(f"  ✅ 批次 {i//batch_size + 1}: 导入 {len(clean_batch)} 条")
        else:
            print(f"  ❌ 批次 {i//batch_size + 1}: HTTP {resp.status_code} - {resp.text[:200]}")

    print(f"\n🎉 导入完成! 共 {total_imported}/{len(customers)} 条")

if __name__ == "__main__":
    main()
