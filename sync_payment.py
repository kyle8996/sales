"""
根据 DK-每月学员缴费明细表.xlsx 同步更新销售系统已报名客户信息
"""
import openpyxl, requests, json
from datetime import datetime

SUPABASE_URL = "https://chasxggorljjqqmficnh.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNoYXN4Z2dvcmxqanFxbWZpY25oIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODE1MDI1NzIsImV4cCI6MjA5NzA3ODU3Mn0.E0xlNCX_jFn8X9LheAR3jFHokzWSkZlXkBDReqTvL6c"
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

def clean_source(raw):
    """清洗客户来源字段"""
    if not raw:
        return None
    raw = str(raw).strip()
    # 异常值修复
    if raw in ['妈妈', '爸爸', '在读二宝']:
        return '在读推荐'
    if raw == '门外海报':
        return 'Walkin'
    if raw == '微信添加':
        return '扫码添加'
    if raw == '在读续费':
        return None  # 续费不改变来源
    if raw == '在读升学':
        return None
    if raw == '在读推荐' or raw == '口碑推荐' or raw == 'Walkin' or raw == '群内添加':
        return raw
    return raw

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    raw = str(val).strip()
    if raw in ['', '/', '—', '无Demo', '年前']:
        return None
    try:
        return datetime.strptime(raw, '%Y-%m-%d').strftime('%Y-%m-%d')
    except:
        return None

def load_payment_data(path):
    """加载所有缴费记录，按名字合并（取最早缴费时间）"""
    wb = openpyxl.load_workbook(path)
    students = {}  # name -> merged data
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).strip()
        
        def get(row, keyword):
            for idx, h in headers.items():
                if keyword in h:
                    return row[idx - 1]
            return None
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            name = get(row, "孩子姓名")
            if not name or str(name).strip() == 'None' or str(name).strip() == '':
                continue
            name = str(name).strip()
            
            birth = get(row, "出生日期")
            phone = get(row, "手机号码")
            source_raw = get(row, "客户来源")
            cls_name = get(row, "报读班级")
            school = get(row, "所在学校")
            demo_date = get(row, "Demo课日期")
            pay_date = get(row, "缴费时间")
            note = get(row, "备注")
            
            birth_str = parse_date(birth)
            phone_str = str(int(phone)) if phone and str(phone).replace('.','').isdigit() else (str(phone).strip() if phone else None)
            source_clean = clean_source(source_raw)
            cls_str = str(cls_name).strip() if cls_name else None
            school_str = str(school).strip() if school and str(school).strip() not in ['', '未上'] else None
            demo_str = parse_date(demo_date)
            pay_str = parse_date(pay_date)
            note_str = str(note).strip() if note and str(note).strip() != 'None' else None
            
            if name in students:
                existing = students[name]
                # 保留最早缴费时间
                if pay_str and (not existing.get('enrollment_date') or pay_str < existing['enrollment_date']):
                    existing['enrollment_date'] = pay_str
                # 保留非空的字段
                for k in ['birth_date','phone','enrolled_class','school','trial_date']:
                    if not existing.get(k):
                        existing[k] = locals()[k if k != 'birth_date' else 'birth_str']
                # 源：优先取非续费/升学的源
                if source_clean and not existing.get('source'):
                    existing['source'] = source_clean
                # 合并备注
                if note_str and note_str != existing.get('note'):
                    existing['note'] = (existing.get('note','') + '; ' + note_str).strip('; ')
            else:
                students[name] = {
                    'name': name,
                    'birth_date': birth_str,
                    'phone': phone_str,
                    'source': source_clean,
                    'enrolled_class': cls_str,
                    'school': school_str,
                    'trial_date': demo_str,
                    'enrollment_date': pay_str,
                    'note': note_str
                }
    
    return students

def main():
    path = "C:/Users/刘新/Desktop/DK·邻里/DK-每月学员缴费明细表.xlsx"
    payment_data = load_payment_data(path)
    
    print(f"📊 缴费表共 {len(payment_data)} 名学员")
    for n, d in sorted(payment_data.items()):
        print(f"  {n} | b:{d.get('birth_date','')} | p:{d.get('phone','')} | src:{d.get('source','')} | cls:{d.get('enrolled_class','')} | school:{d.get('school','')} | trial:{d.get('trial_date','')} | enroll:{d.get('enrollment_date','')}")
    
    # 获取现有客户
    print("\n📡 获取现有客户...")
    r = requests.get(f"{SUPABASE_URL}/rest/v1/customers?select=*", headers={**HEADERS, "Accept": "application/json"})
    existing = r.json()
    print(f"  现有 {len(existing)} 条")
    
    # 构建查找映射 (name -> customer)
    name_map = {}
    for c in existing:
        n = c['name']
        if n not in name_map:
            name_map[n] = []
        name_map[n].append(c)
    
    updated = 0
    created = 0
    skipped = 0
    
    for name, pd in payment_data.items():
        matches = name_map.get(name, [])
        
        if matches:
            c = matches[0]
            cid = c['id']
            
            # 构建更新数据
            update = {}
            if pd.get('birth_date') and not c.get('birth_date'):
                update['birth_date'] = pd['birth_date']
            if pd.get('enrolled_class') and not c.get('enrolled_class'):
                update['enrolled_class'] = pd['enrolled_class']
            if pd.get('enrollment_date') and not c.get('enrollment_date'):
                update['enrollment_date'] = pd['enrollment_date']
            if pd.get('school') and not c.get('school'):
                update['school'] = pd['school']
            if pd.get('source') and pd['source'] not in (c.get('source') or ''):
                # source 有更精确的信息才更新
                update['source'] = pd['source']
            if pd.get('trial_date') and not c.get('trial_date'):
                update['trial_date'] = pd['trial_date'] + 'T00:00'
            
            # 确保状态为已报名
            if c.get('status') != '已报名':
                update['status'] = '已报名'
            
            if update:
                update['updated_at'] = datetime.now().isoformat()
                resp = requests.patch(
                    f"{SUPABASE_URL}/rest/v1/customers?id=eq.{cid}",
                    headers=HEADERS,
                    json=update
                )
                if resp.status_code in [200, 201, 204]:
                    updated += 1
                    print(f"  ✅ 更新: {name} -> {list(update.keys())}")
                else:
                    print(f"  ❌ 更新失败 {name}: {resp.status_code} {resp.text[:100]}")
            else:
                skipped += 1
                print(f"  ⏭️  跳过: {name} (数据已完整)")
        else:
            # 不存在，创建新客户
            new_cust = {
                'name': name,
                'birth_date': pd.get('birth_date'),
                'phone': pd.get('phone'),
                'source': pd.get('source'),
                'enrolled_class': pd.get('enrolled_class'),
                'school': pd.get('school'),
                'trial_date': (pd['trial_date'] + 'T00:00') if pd.get('trial_date') else None,
                'enrollment_date': pd.get('enrollment_date'),
                'status': '已报名',
                'owner': 'Kyle',
                'level': 'B',
                'notes': pd.get('note'),
                'created_at': datetime.now().isoformat()
            }
            resp = requests.post(f"{SUPABASE_URL}/rest/v1/customers", headers=HEADERS, json=new_cust)
            if resp.status_code in [200, 201, 204]:
                created += 1
                print(f"  🆕 新增: {name}")
            else:
                print(f"  ❌ 新增失败 {name}: {resp.status_code} {resp.text[:200]}")
    
    print(f"\n🎉 同步完成! 新增 {created} | 更新 {updated} | 跳过 {skipped}")

if __name__ == "__main__":
    main()
